from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
from sqlalchemy import func
from app.database import SessionLocal
from app.models.contabilita import MovimentoContabile
from app.schemas.pagamenti import (
    MovimentoContabileCreate, MovimentoContabileRead,
    RendicontoResponse, RendicontoRigaCategoria, RendicontoRigaMensile,
)
from typing import List, Optional
from datetime import date

router = APIRouter(prefix="/prima-nota", tags=["Contabilità"])

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


@router.get("/", response_model=List[MovimentoContabileRead])
def lista_movimenti(
    data_da: Optional[date] = None,
    data_a: Optional[date] = None,
    tipo: Optional[str] = None,
    categoria: Optional[str] = None,
    db: Session = Depends(get_db),
):
    q = db.query(MovimentoContabile)
    if data_da:
        q = q.filter(MovimentoContabile.data >= data_da)
    if data_a:
        q = q.filter(MovimentoContabile.data <= data_a)
    if tipo:
        q = q.filter(MovimentoContabile.tipo == tipo)
    if categoria:
        q = q.filter(MovimentoContabile.categoria == categoria)
    return q.order_by(MovimentoContabile.data.desc(), MovimentoContabile.id.desc()).all()


@router.post("/", response_model=MovimentoContabileRead)
def crea_movimento(movimento: MovimentoContabileCreate, db: Session = Depends(get_db)):
    db_movimento = MovimentoContabile(**movimento.model_dump())
    db.add(db_movimento)
    db.commit()
    db.refresh(db_movimento)
    return db_movimento


@router.put("/{movimento_id}", response_model=MovimentoContabileRead)
def aggiorna_movimento(movimento_id: int, movimento: MovimentoContabileCreate, db: Session = Depends(get_db)):
    db_movimento = db.query(MovimentoContabile).filter(MovimentoContabile.id == movimento_id).first()
    if not db_movimento:
        raise HTTPException(status_code=404, detail="Movimento non trovato")
    if db_movimento.pagamento_id:
        raise HTTPException(
            status_code=400,
            detail="Questo movimento è generato automaticamente da un incasso e non può essere modificato manualmente."
        )
    for key, value in movimento.model_dump().items():
        setattr(db_movimento, key, value)
    db.commit()
    db.refresh(db_movimento)
    return db_movimento


@router.delete("/{movimento_id}")
def elimina_movimento(movimento_id: int, db: Session = Depends(get_db)):
    db_movimento = db.query(MovimentoContabile).filter(MovimentoContabile.id == movimento_id).first()
    if not db_movimento:
        raise HTTPException(status_code=404, detail="Movimento non trovato")
    if db_movimento.pagamento_id:
        raise HTTPException(
            status_code=400,
            detail="Questo movimento è generato automaticamente da un incasso: elimina/modifica il pagamento collegato invece."
        )
    db.delete(db_movimento)
    db.commit()
    return {"messaggio": "Movimento eliminato"}


# ---- RENDICONTO ECONOMICO ----

rendiconto_router = APIRouter(prefix="/rendiconto", tags=["Contabilità"])


@rendiconto_router.get("/", response_model=RendicontoResponse)
def rendiconto(
    data_da: Optional[date] = None,
    data_a: Optional[date] = None,
    db: Session = Depends(get_db),
):
    oggi = date.today()
    if not data_da:
        data_da = date(oggi.year, 1, 1)
    if not data_a:
        data_a = date(oggi.year, 12, 31)

    movimenti = db.query(MovimentoContabile).filter(
        MovimentoContabile.data >= data_da, MovimentoContabile.data <= data_a
    ).all()

    totale_entrate = sum(float(m.importo) for m in movimenti if m.tipo == "entrata" or getattr(m.tipo, "value", m.tipo) == "entrata")
    totale_uscite = sum(float(m.importo) for m in movimenti if m.tipo == "uscita" or getattr(m.tipo, "value", m.tipo) == "uscita")

    # Aggregazione per categoria
    per_categoria_map = {}
    for m in movimenti:
        cat = m.categoria or "Senza categoria"
        tipo_val = getattr(m.tipo, "value", m.tipo)
        if cat not in per_categoria_map:
            per_categoria_map[cat] = {"entrate": 0.0, "uscite": 0.0}
        per_categoria_map[cat]["entrate" if tipo_val == "entrata" else "uscite"] += float(m.importo)

    per_categoria = [
        RendicontoRigaCategoria(categoria=cat, entrate=v["entrate"], uscite=v["uscite"])
        for cat, v in sorted(per_categoria_map.items(), key=lambda x: -(x[1]["entrate"] + x[1]["uscite"]))
    ]

    # Aggregazione per mese
    per_mese_map = {}
    for m in movimenti:
        chiave = m.data.strftime("%Y-%m")
        tipo_val = getattr(m.tipo, "value", m.tipo)
        if chiave not in per_mese_map:
            per_mese_map[chiave] = {"entrate": 0.0, "uscite": 0.0}
        per_mese_map[chiave]["entrate" if tipo_val == "entrata" else "uscite"] += float(m.importo)

    per_mese = [
        RendicontoRigaMensile(mese=k, entrate=v["entrate"], uscite=v["uscite"], saldo=v["entrate"] - v["uscite"])
        for k, v in sorted(per_mese_map.items())
    ]

    return RendicontoResponse(
        periodo_da=data_da,
        periodo_a=data_a,
        totale_entrate=totale_entrate,
        totale_uscite=totale_uscite,
        saldo=totale_entrate - totale_uscite,
        per_categoria=per_categoria,
        per_mese=per_mese,
    )


@router.get("/export/excel")
def esporta_prima_nota_excel(
    data_da: Optional[date] = None,
    data_a: Optional[date] = None,
    db: Session = Depends(get_db),
):
    """Esporta i movimenti di Prima Nota in Excel nello stesso formato tradizionale
    usato finora dall'associazione: Data, Descrizione, Entrata, Uscita, Saldo (progressivo)."""
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    import io as _io

    q = db.query(MovimentoContabile).order_by(MovimentoContabile.data, MovimentoContabile.id)
    if data_da:
        q = q.filter(MovimentoContabile.data >= data_da)
    if data_a:
        q = q.filter(MovimentoContabile.data <= data_a)
    movimenti = q.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Cassa"

    ws.merge_cells('A1:E1')
    ws['A1'] = "Cassa"
    ws['A1'].font = Font(bold=True, size=12)
    ws['A1'].alignment = Alignment(horizontal='center')

    intestazioni = ['Data', 'Descrizione', 'Entrata', 'Uscita', 'Saldo']
    for col, testo in enumerate(intestazioni, start=1):
        cell = ws.cell(row=2, column=col, value=testo)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="1E3A8A")
        cell.alignment = Alignment(horizontal='center')

    saldo = 0.0
    for i, m in enumerate(movimenti, start=3):
        tipo_val = getattr(m.tipo, "value", m.tipo)
        entrata = float(m.importo) if tipo_val == "entrata" else None
        uscita = float(m.importo) if tipo_val == "uscita" else None
        saldo += (entrata or 0) - (uscita or 0)
        ws.cell(row=i, column=1, value=m.data.strftime('%d/%m/%Y'))
        ws.cell(row=i, column=2, value=m.descrizione)
        ws.cell(row=i, column=3, value=entrata)
        ws.cell(row=i, column=4, value=uscita)
        ws.cell(row=i, column=5, value=round(saldo, 2))

    for col, larghezza in zip(range(1, 6), [14, 45, 14, 14, 14]):
        ws.column_dimensions[get_column_letter(col)].width = larghezza
    ws.freeze_panes = "A3"

    buffer = _io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=prima_nota.xlsx"}
    )
